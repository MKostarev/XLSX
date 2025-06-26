import openpyxl
from celery import shared_task

from upload_files.models import ProductGroup, Product


@shared_task(bind=True)
def process_excel_file(self, file_path):
    try:
        wb = openpyxl.load_workbook(filename=file_path)
        sheet = wb.active

        # Получаем или создаем дефолтную группу "Запчасти"
        default_group, _ = ProductGroup.objects.get_or_create(name='Запчасти')

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Предполагаем следующую структуру столбцов:
                # 0 - brand, 1 - article, 2 - trading_numbers, 3 - description,
                # 4 - additional_name, 5 - product_group_name, 6 - product_status,
                # 7 - specifications

                if not row[0] or not row[1]:  # Проверка обязательных полей
                    continue

                # Обрабатываем товарную группу
                group_name = row[5] if row[5] else 'Запчасти'
                product_group, _ = ProductGroup.objects.get_or_create(
                    name=group_name,
                    defaults={'parent': default_group}
                )

                # Создаем/обновляем товар
                Product.objects.update_or_create(
                    article=row[1],  # Уникальный артикул
                    defaults={
                        'brand': row[0],
                        'trading_numbers': row[2] if len(row) > 2 else '',
                        'description': row[3] if len(row) > 3 else '',
                        'additional_name': row[4] if len(row) > 4 else '',
                        'product_group': product_group,
                        'product_status': row[6] if len(row) > 6 else 'Новый',
                        'specifications': row[7] if len(row) > 7 else ''
                    }
                )

            except Exception as e:
                print(f"Ошибка в строке {row_idx}: {e}")
                continue

        return f"Обработано {sheet.max_row - 1} строк"

    except Exception as e:
        raise self.retry(exc=e, countdown=60)